# this an example for pull request
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
import os

# Create workbook and sheet 
if not os.path.exists("student_scores.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Score Tracker"
    ws.append(["Name", "Score", "Remarks"])
    wb.save("student_scores.xlsx")

def validate_inputs():
    name = name_entry.get()
    score = score_entry.get()

    if not name or not score:
        messagebox.showerror("Input Error", "All fields are required!")
        return False
    try:
        int(score)
    except ValueError:
        messagebox.showerror("Input Error", "Score must be a number!")
        return False
    return True

def save_to_excel():
    if not validate_inputs():
        return

    name = name_entry.get()
    score = int(score_entry.get())
    status = "Pass" if score >= 75 else "Fail"

    wb = load_workbook("student_scores.xlsx")
    ws = wb["Student Score Tracker"]

    # Remove existing average row if any
    if ws.max_row > 1:
        last_row = ws.max_row
        if ws.cell(row=last_row, column=1).value == "Average":
            ws.delete_rows(last_row)

    ws.append([name, score, status])

    # Calculate average
    scores = [cell.value for cell in ws["B"][1:] if isinstance(cell.value, (int, float))]
    average_score = sum(scores) / len(scores) if scores else 0
    ws.append(["Average", average_score, ""])

    format_excel(ws)
    wb.save("student_scores.xlsx")

    messagebox.showinfo("Success", "Data saved successfully!")

    name_entry.delete(0, tk.END)
    score_entry.delete(0, tk.END)

def format_excel(ws: Worksheet):
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

def show_data():
    wb = load_workbook("student_scores.xlsx")
    ws = wb["Student Score Tracker"]

    data_window = tk.Toplevel(window)
    data_window.title("Stored Student Scores")

    col_widths = [30, 15, 15]  # Wider Name column

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        for j, value in enumerate(row):
            width = col_widths[j] if j < len(col_widths) else 15

            is_header = i == 0
            is_remarks_col = j == 2
            is_average_row = row[0] == "Average"
            is_pass_fail = is_remarks_col and isinstance(value, str) and value in ["Pass", "Fail"]

            font_style = ("Arial", 10, "bold") if is_header or is_average_row or is_pass_fail else ("Arial", 10)

            label = tk.Label(
                data_window,
                text=value,
                width=width,
                font=font_style,
                borderwidth=1,
                relief="solid",
                padx=6,
                pady=3,
                anchor="w" if j == 0 else "center"
            )
            label.grid(row=i, column=j)

# GUI Setup
window = tk.Tk()
window.title("Student Score Tracker")

tk.Label(window, text="Name").grid(row=0, column=0, padx=10, pady=5, sticky="w")
tk.Label(window, text="Score").grid(row=1, column=0, padx=10, pady=5, sticky="w")

name_entry = tk.Entry(window, width=40)
score_entry = tk.Entry(window, width=40)

name_entry.grid(row=0, column=1, padx=20, pady=5)
score_entry.grid(row=1, column=1, padx=20, pady=5)

tk.Button(window, text="Submit", command=save_to_excel, width=20, bg="#C8A2C8", fg="white").grid(row=3, column=0, columnspan=2, pady=10)
tk.Button(window, text="View Stored Data", command=show_data, width=20, bg="#965A96", fg="white").grid(row=4, column=0, columnspan=2)

window.mainloop()

